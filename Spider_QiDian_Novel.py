#coding=utf8
_Author_ = 'Alvis'
_Date_ = '2018-04-29 20:56'

import re
import requests
import openpyxl
from time import sleep
import time
from requests.packages import urllib3
class Get_qd_novel:

    def __init__(self):
        self.novel_name = []
        self.novel_author = []
        self.novel_big_type = []
        self.novel_small_type = []
        self.novel_status=[]
        self.novel_intro=[]
        with open("qiandian.log", 'a+') as f:
            f.write("程序启动时间为:"+time.strftime("%Y-%m-%d %H:%M:%S")  + "\n")
            f.close()

    def GetNovelData(self):
        for page in range(1,44974):
            headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36"}
            qd_url = 'https://www.qidian.com/all?orderId=&page=%s&style=1&pageSize=20&siteid=1&pubflag=0&hiddenField=0'%page
            try:
                response = requests.get(qd_url,headers=headers,verify=False,timeout=5)
            except Exception as net_error:
                with open("qiandian.log",'a+') as f:
                    f.write(time.strftime("%Y-%m-%d %H:%M:%S  ")+"第%s页无法获取到数据  "%page +"错误为:%s"%net_error + "\n")
                    f.close()
                continue
            urllib3.disable_warnings()
            html = response.text

            # 小说名称
            name_contents = re.findall('<h4><a href="//book.qidian.com/info/(.*?)/a></h4>', html, re.S)
            # 小说作者
            author_contents = re.findall('<a class="name" href="//my.qidian.com/author/(.*?)/a>', html, re.S)
            # 小说大类型
            big_type_contents = re.findall('</em><a href="//www.(.*?)/a>', html, re.S)
            # 小说小类型
            small_type_contents = re.findall('data-eid="qd_B61">(.*?)</a>', html, re.S)
            # 小说状态
            status_contents = re.findall('</em><span >(.*?)</span>', html, re.S)
            # 小说简介
            intro_contents = re.findall('<p class="intro">(.*?)</p>', html, re.S)

            for i in range(len(intro_contents)):
                self.novel_name.append(re.findall('>(.*?)<', name_contents[i], re.S)[0])
                self.novel_author.append(re.findall('>(.*?)<', author_contents[i], re.S)[0])
                self.novel_big_type.append(re.findall('>(.*?)<', big_type_contents[i], re.S)[0])
                self.novel_small_type.append(small_type_contents[i])
                self.novel_status.append(status_contents[i])
                self.novel_intro.append(intro_contents[i].strip())

            print("已爬取第%s页内容"%page)

    def SaveDataToExcel(self):
        with open("qiandian.log", 'a+') as f:
            f.write("正在写入数据  "+time.strftime("%Y-%m-%d %H:%M:%S")  + "\n")
            f.close()
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename="起点中文网小说名单.xlsx")
        except Exception as e:
            with open("qiandian.log", 'a+') as f:
                f.write("文件名不存在,进行创建  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                f.close()
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title ='小说信息'
        ws.cell(row=1, column=1, value='小说名称')
        ws.cell(row=1, column=2, value='小说作者')
        ws.cell(row=1, column=3, value='小说大类型')
        ws.cell(row=1, column=4, value='小说小类型')
        ws.cell(row=1, column=5, value='小说状态')
        ws.cell(row=1, column=6, value='小说简介')

        with open("qiandian.log", 'a+') as f:
            f.write("开始写入小说名  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        #统计写入小说名数量
        name_num=0
        for i in range(len(self.novel_name)):
            try:
                ws.cell(row=i + 2, column=1, value=self.novel_name[i])
                name_num=name_num+1
            except Exception as name_error:
                with open("qiandian.log", 'a+') as f:
                    f.write("无法写入第%s本小说:%s,原因:%s  "%(i,self.novel_name[i],name_error) + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        with open("qiandian.log", 'a+') as f:
            f.write("开始写入小说作者  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        # 统计写入小说作者数量
        author_num = 0
        for i in range(len(self.novel_author)):
            try:
                ws.cell(row=i + 2, column=2, value=self.novel_author[i])
                author_num=author_num+1
            except Exception as author_error:
                with open("qiandian.log", 'a+') as f:
                    f.write("无法写入第%s个作者:%s,原因:%s  "%(i,self.novel_author[i],author_error) + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        with open("qiandian.log", 'a+') as f:
            f.write("开始写入小说大类型  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        # 统计写入小说大类型数量
        big_type_num = 0
        for i in range(len(self.novel_big_type)):
            try:
                ws.cell(row=i + 2, column=3, value=self.novel_big_type[i])
                big_type_num=big_type_num+1
            except Exception as big_type_error:
                with open("qiandian.log", 'a+') as f:
                    f.write("无法写入第%s个大分类:%s,原因:%s  " % (i, self.novel_big_type[i],big_type_error) + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        with open("qiandian.log", 'a+') as f:
            f.write("开始写入小说小类型  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        # 统计写入小说小类型数量
        small_type_num = 0
        for i in range(len(self.novel_small_type)):
            try:
                ws.cell(row=i + 2, column=4, value=self.novel_small_type[i])
                small_type_num=small_type_num+1
            except Exception as small_type_error:
                with open("qiandian.log", 'a+') as f:
                    f.write("无法写入第%s个小分类:%s,原因:%s "% (i, self.novel_small_type[i],small_type_error) + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        with open("qiandian.log", 'a+') as f:
            f.write("开始写入小说状态  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        # 统计写入小说状态数量
        status_num = 0
        for i in range(len(self.novel_status)):
            try:
                ws.cell(row=i + 2, column=5, value=self.novel_status[i])
                status_num=status_num+1
            except Exception as status_error:
                with open("qiandian.log", 'a+') as f:
                    f.write("无法写入第%s个状态:%s,原因:%s  " % (i, self.novel_status[i],status_error) + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        with open("qiandian.log", 'a+') as f:
            f.write("开始写入小说简介  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        # 统计写入小说简介数量
        intro_num = 0
        for i in range(len(self.novel_intro)):
            try:
                ws.cell(row=i + 2, column=6, value=self.novel_intro[i])
                intro_num=intro_num+1
            except Exception as intro_error:
                with open("qiandian.log", 'a+') as f:
                    f.write("无法写入第%s个简介:%s,原因:%s  " % (i, self.novel_intro[i],intro_error) + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        with open("qiandian.log", 'a+') as f:
            f.write(time.strftime("%Y-%m-%d %H:%M:%S")+"  当前总共爬取%s本小说，写入%s本小说;当前总共爬取%s个作者，写入%s个作者; 当前总共爬取%s个大分类,写入%s个大分类; 当前总共爬取%s个小分类,写入%s个小分类; 当前总共爬取%s个状态,写入%s个状态; 当前总共爬取%s个简介,写入%s个简介"
                    %(len(self.novel_name),name_num,len(self.novel_author),author_num,len(self.novel_big_type),big_type_num,len(self.novel_small_type),small_type_num,len(self.novel_status),status_num,len(self.novel_intro),intro_num)
                     + "\n")


        try:
            with open("qiandian.log", 'a+') as f:
                f.write("正在保存文件  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            wb.save("起点中文网小说名单.xlsx")
        except Exception as save_fault:
            with open("qiandian.log", 'a+') as f:
                f.write("文件保存失败,原因: %s  "%save_fault + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        with open("qiandian.log", 'a+') as f:
            f.write("程序结束时间为:"+time.strftime("%Y-%m-%d %H:%M:%S")  + "\n")
            f.close()



if __name__=="__main__":
    Spider=Get_qd_novel()
    Spider.GetNovelData()
    Spider.SaveDataToExcel()