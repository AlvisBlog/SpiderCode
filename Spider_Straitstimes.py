# -*- coding: utf-8 -*-

import requests
import re
from requests.packages import urllib3
from time import sleep
import openpyxl
class Spider_News:

    def __init__(self):
        self.news=[]
        self.all_topic_url=[]
        self.all_news_url=[]


    def GetTopicUrl(self):

        print("——————————开始获取所有的topic链接——————————")

        #所有的父topic
        all_big_topic_url=['/singapore','/politics','/asia','/world','/videos','/multimedia','/lifestyle','/lifestyle/food','/forum','/opinion','/business','/sport','/tech']

        #以下为提取每个父topic下的子topic的过程
        for j in range(len(all_big_topic_url)):

            print("开始获取第%s个父topic下的所有子topic"%(j+1))

            try:
                # 忽略警告信息
                urllib3.disable_warnings()

                #拼接父topic的url,进行请求
                response=requests.get("https://www.straitstimes.com"+all_big_topic_url[j],verify=False)

                #获取父topic页面的源码
                html=response.text

                #获取子topic,形式为:/courts-crime
                url=re.findall('<li class=".*?leaf"><a href="%s(.*?)">'%all_big_topic_url[j],html,re.S)

                #因为有部分父topic下不存在子topic需要做判断处理,即不存在子topic时,则父topic为子topic
                if len(url)==0:

                    self.all_topic_url.append((all_big_topic_url[j]))

                    print("该父topic:%s下没有子topic" % (all_big_topic_url[j]))

                else:

                    # 对子topic进行拼接,形式:/singapore/courts-crime
                    for i in range(len(url)):

                        #替换部分不规则脏数据
                        part_topic_url=all_big_topic_url[j]+url[i].replace('" title="',"")

                        #添加入所有子topic列表
                        self.all_topic_url.append((part_topic_url))

                        print("已获取该父topic:%s的%s个子topic"%(all_big_topic_url[j],i+1))

            except Exception as e:

                print("无法访问该网址:%s,错误:%s"%(all_big_topic_url[j],e))

                #若当前父topic无法访问,则进行放弃,继续访问下一父topic
                continue

            #约定每5秒访问下一个父topic
            sleep(5)
            print("\n")


        return self.all_topic_url


    def GetNewsUrl(self,max_page):
        '''max_page为每个子topic下需要翻过的页数,定义后,每个子topic将按此约定进行访问'''
        print("——————————开始获取新闻链接——————————")
        sleep(3)

        #遍历每个topic的url
        for url in self.all_topic_url:

            #约定每个topic下的页数
            for page in range(0,max_page+1):

                print("正在获取链接:%s的第%s页的新闻链接"%("https://www.straitstimes.com"+url,page))

                #头部信息
                headers={'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36'}

                # 忽略警告信息
                urllib3.disable_warnings()

                try:

                    #首页page为0，替换为空
                    if page!=0:
                        response=requests.get("https://www.straitstimes.com"+url+'?page=%s'%page,verify=False,headers=headers)

                    else:
                        response = requests.get("https://www.straitstimes.com"+url)

                    #获取网页源码
                    html=response.text

                    #提取url
                    url_news=re.findall('%s/(.*?)">'%url,html,re.S)

                    for i in range(len(url_news)):

                        self.all_news_url.append("https://www.straitstimes.com"+url+"/"+url_news[i])

                        print("已获取%s条新闻链接"%len(self.all_news_url))

                    #网络原因,每10秒访问下一页,视情况而定
                    sleep(10)

                    print("\n")

                except Exception as e:
                    print("访问出错:%s"%e)



        return self.all_news_url


    def GetNewsContent(self):
        print("——————————开始获取新闻内容——————————")
        sleep(3)
        try:
            for url in self.all_news_url:
                # url='https://www.straitstimes.com/lifestyle/entertainment/angelababy-responds-to-criticism-for-special-treatment-on-keep-running'
                response=requests.get(url)
                html=response.text

                #获取正文信息
                try:
                    data=re.findall('<p>(.*?)</p>',html,re.S)
                    if data==[]:
                        article='无正文信息'
                    else:
                        article = ''
                        for i in range(1, len(data) - 2):
                            article = article + data[i]
                except Exception as e1:
                    article="无正文信息,数据获取失败"

                #获取标题信息
                try:
                    title=re.findall('<title>(.*?)</title>',html,re.S)[0]
                    if title=='':
                        title="无标题信息"
                except Exception as e2:
                    title="无标题信息,数据获取失败"

                #获取发布时间
                try:
                    pubdate=re.findall('"pubdate":"(.*?)",',html,re.S)[0]
                    if pubdate=='':
                        pubdate='无更新时间信息'
                except Exception as e3:
                    pubdate="无更新时间信息,数据获取失败"

                #获取作者
                try:
                    author=re.findall('"author": "(.*?)",',html,re.S)[0].replace("+"," ")
                    if author=="":
                        author='无作者信息'
                except Exception as e4:
                    author="无作者信息,该数据获取失败"

                self.news.append({'title':title,'article':article,'pubdate':pubdate,'author':author})

                print("已获取%s条新闻信息"%len(self.news))

                #每10秒访问下一个新闻页,时间视情况而定
                sleep(1)

                print("\n")

        except Exception as net_error:
            print("网络错误,%s:"%net_error)



        return self.news


    def SaveData(self):
        print("——————————开始保存数据——————————")
        sleep(3)
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename='数据信息.xlsx')
        except Exception as e:
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        print(wb.sheetnames)
        # 为新表命名
        ws.title ='news'
        ws.cell(row=1, column=1, value="新闻标题")
        ws.cell(row=1, column=2, value="新闻正文")
        ws.cell(row=1, column=3, value="新闻发布时间")
        ws.cell(row=1, column=4, value="新闻作者")
        try:
            for i in range(len(self.news)):
                ws.cell(row=i + 1, column=1, value=self.news[1]['title'])
                ws.cell(row=i + 1, column=2, value=self.news[1]['article'])
                ws.cell(row=i + 1, column=3, value=self.news[1]['pubdate'])
                ws.cell(row=i + 1, column=4, value=self.news[1]['author'])
        except Exception as e:
            print(e)

        wb.save("数据信息.xlsx")

if __name__ == '__main__':
    Spider=Spider_News()
    Spider.GetTopicUrl()
    Spider.GetNewsUrl(10)
    Spider.GetNewsContent()
    Spider.SaveData()