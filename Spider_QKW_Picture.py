# -*- coding: UTF-8 -*-

_Author_ = 'Alvis'

_Date_ = '2018/4/14 14:41'

import requests
import re
import xlwt
import os
import time
import openpyxl

class Get_qkw_picture:

    def __init__(self):
        self.Pic_links=[]

    #获取图片链接地址
    def Get_Picture_Links(self):
        #url地址
        url="http://588ku.com/tuku/keji.html"
        #发送Http请求,并接收返回
        response=requests.get(url)
        #获取网页源代码
        html=response.text
        self.Pic_links=re.findall('original="(.*?)!/fh',html)
        return self.Pic_links

    #下载图片至目录image
    def Download_Pic_To_Dir(self):
        try:
            os.mkdir('./image')
        except Exception as e:
            pass
        for i in range(len(self.Pic_links)):
            r=requests.get(self.Pic_links[i])
            with open('./image/%s_%s.jpg'%(i,time.time()),'wb') as f:
                f.write(r.content)

    #下载链接保存到xls文件
    def Save_PicLink_To_XLS(self,ExcelName,SheetName):
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename='%s.xlsx' % ExcelName)
        except Exception as e:
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.get_sheet_names()
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            wb.remove_sheet(wb.get_sheet_by_name(name))
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title = SheetName
        ws.cell(row=1, column=1, value='图片链接')
        for i in range(len(self.Pic_links)):
            ws.cell(row=2+i, column=1, value=self.Pic_links[i])
        wb.save('%s.xlsx'%ExcelName)

if __name__=="__main__":
   Spider=Get_qkw_picture()
   Spider.Get_Picture_Links()
   Spider.Save_PicLink_To_XLS("图片集","千库网图片")
   Spider.Download_Pic_To_Dir()