# -*- coding: UTF-8 -*-
_Author_ = 'Alvis'
_Date_ = '21:51 2018-03-26'
_README_='淘宝信息爬取'

import re
import requests
import json
import time
import openpyxl

class TaoBao():
    def __init__(self):
        self.DATA=[]
        self.goods_name='软件测试'

    def GetData(self,page):

        #url地址
        url="https://s.taobao.com/search?q={}&imgfile=&js=1&stats_click=search_radio_all%3A1&initiative_id=staobaoz_20180326&ie=utf8".format(self.goods_name)

        #发送Http请求,并接收返回
        response=requests.get(url)

        #html源码
        html=response.text

        content=re.findall(u'g_page_config = (.*?)g_srp_loadCss',html,re.S)[0].strip()[:-1]


        #格式化json
        content=json.loads(content)

        #获取商品信息列表，商品名称，标题，标价，购买人数，是否包邮，是否天猫，地区，店名，url
        data_list=content['mods']['itemlist']['data']['auctions']

        #提取数据，分析content结构
        for item in data_list:

            temp={
                'title':item['title'],
                'view_price': item['view_price'],
                'view_sales': item['view_sales'],
                'view_fee': '否' if float(item['view_fee']) else '是',
                'isTmall': '是' if item['shopcard']['isTmall'] else '否',
                'area': item['item_loc'],
                'name': item['nick'],
                'detail_url': item['detail_url'],
            }
        #每次获取完数据就加入self.DATA列表里面
            self.DATA.append(temp)

        #需要cookie保持,验证，将第一次response的cookies传给第二次
        cookies=response.cookies

        url2="https://s.taobao.com/api?_ksTS=1522072777204_224&callback=jsonp225&ajax=true&m=customized&stats_click=search_radio_all:1&q={}&s=36&imgfile=&initiative_id=staobaoz_20180326&bcoffset=0&js=1&ie=utf8&rn=8c0acb1785ec2c68d3a3e9a3d532d29a".format(self.goods_name)

        response2=requests.get(url2,cookies=cookies)

        html2=response2.text

        content=re.findall(r'{.*}',html2)[0]

        #格式化json
        content=json.loads(content)

        data_list=content['API.CustomizedApi']['itemlist']['auctions']

        for item in data_list:

            temp={
                'title':item['title'],
                'view_price': item['view_price'],
                'view_sales': item['view_sales'],
                'view_fee': '否' if float(item['view_fee']) else '是',
                'isTmall': '是' if item['shopcard']['isTmall'] else '否',
                'area': item['item_loc'],
                'name': item['nick'],
                'detail_url': item['detail_url'],
            }
        #每次获取完数据就加入self.DATA列表里面
            self.DATA.append(temp)

        #将cookies保持下来
        cookies=response2.cookies

        #翻页获取数据
        for i in range(1,page):
            ktsts=time.time()
            _ksTS='%s_%s' % (int(ktsts*1000),str(ktsts)[-3:])
            callback="jsonp%s" % (int(str(ktsts)[-3:]) + 1)
            data_value=44 * i
            #翻页逻辑
            url="https://s.taobao.com/search?self.DATA-key=s&self.DATA-value={}&ajax=true&_ksTS={}&callback={}&q={}&imgfile=&js=1&stats_click=search_radio_all%3A1&initiative_id=staobaoz_20180327&ie=utf8&bcoffset=4&ntoffset=0&p4ppushleft=1%2C48".format(data_value,_ksTS,callback,self.goods_name)
            #cookies保持
            response3=requests.get(url,cookies=cookies)
            html=response3.text
            data_list = content['API.CustomizedApi']['itemlist']['auctions']

            for item in data_list:
                temp = {
                    'title': item['title'],
                    'view_price': item['view_price'],
                    'view_sales': item['view_sales'],
                    'view_fee': '否' if float(item['view_fee']) else '是',
                    'isTmall': '是' if item['shopcard']['isTmall'] else '否',
                    'area': item['item_loc'],
                    'name': item['nick'],
                    'detail_url': item['detail_url'],
                }
                # 每次获取完数据就加入self.DATA列表里面，根据翻页数最终得到self.DATA的数目
                self.DATA.append(temp)

    def SaveData(self):
        #持久化，写入excel文件
        try:
            wb=openpyxl.load_workbook(filename="淘宝商品数据.xlsx")
        except Exception as e:
            wb=openpyxl.Workbook()
        all_sheet=wb.sheetnames
        if 'Sheet' in all_sheet:
            del wb['Sheet']
        ws=wb.create_sheet()
        ws.title=self.goods_name
        ws.cell(row=1,column=1,value='标题')
        ws.cell(row=1,column=2,value='标价')
        ws.cell(row=1,column=3,value='购买人数')
        ws.cell(row=1,column=4,value='是否包邮')
        ws.cell(row=1,column=5,value='是否天猫')
        ws.cell(row=1,column=6,value='地区')
        ws.cell(row=1,column=7,value='店名')
        ws.cell(row=1,column=8,value='URL')

        #写内容
        for i in range(len(self.DATA)):
            ws.cell(row=i+2, column=1, value=self.DATA[i]['title'].replace("<span class=H>"," ").replace("</span>",""))
            ws.cell(row=i+2, column=2, value=self.DATA[i]['view_price'])
            ws.cell(row=i+2, column=3, value=self.DATA[i]['view_sales'])
            ws.cell(row=i+2, column=4, value=self.DATA[i]['view_fee'])
            ws.cell(row=i+2, column=5, value=self.DATA[i]['isTmall'])
            ws.cell(row=i+2, column=6, value=self.DATA[i]['area'])
            ws.cell(row=i+2, column=7, value=self.DATA[i]['name'])
            ws.cell(row=i+2, column=8, value=self.DATA[i]['detail_url'])

        wb.save('淘宝商品数据.xlsx')

if __name__=="__main__":
    Spider=TaoBao()
    Spider.GetData(10)
    Spider.SaveData()