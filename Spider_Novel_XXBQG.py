#coding=utf8
_Author_ = 'Alvis'
_Date_ = '2018-05-02 19:21'

import requests
import re
import openpyxl
from requests.packages import urllib3
from time import sleep

class Spider_bqg_novel:

    #初始化数据
    def __init__(self):
        #章节名列表
        self.chapter_name=[]
        #章节链接列表
        self.chapter_url=[]
        #章节内容
        self.chapter_content=[]
        #小说作者
        self.author=[]
        #小说状态
        self.status=[]
        #小说最后更新时间
        self.last_update_time=[]
        #小说最近更新的章节名称
        self.last_chapter_name=[]
        self.chapter_num=1

    #获取章节名称及链接、内容,需要传入小说页面链接
    def GetChapterData(self,url):
        # 返回信息
        response = requests.get(url,verify=False)
        urllib3.disable_warnings()
        # 源网页代码
        html = response.text.encode('ISO-8859-1').decode('utf8')
        #小说状态
        self.status=re.findall('<meta property="og:novel:status" content="(.*?)"/>',html,re.S)
        #获取作者
        self.author=re.findall('<meta property="og:novel:author" content="(.*?)"/>',html,re.S)
        #获取最后更新时间
        self.last_update_time=re.findall('<meta property="og:novel:update_time" content="(.*?)"/>',html,re.S)
        #获取最近更新的章节名称
        self.last_chapter_name=re.findall('<meta property="og:novel:latest_chapter_name" content="(.*?)"/>',html,re.S)
        # 网页中包含章节名称及链接的数据
        content = re.findall('<dl>(.*?)</dl>', html, re.S)
        # 获取章节部分链接，缺少"https://www.xxbiquge.com"
        chapter_url_content = re.findall('<dd><a href="(.*?)">', content[0], re.S)
        # 获取章节名称
        print("正在获取章节名称")
        self.chapter_name = re.findall('">(.*?)</a>', content[0], re.S)
        print("已获取章节名称")
        # 获取章节链接，补全"https://www.xxbiquge.com"
        print("正在获取章节链接")
        for link in chapter_url_content:
            self.chapter_url.append("https://www.xxbiquge.com" + link)
        print("已获取章节链接")
        # 获取章节内容
        print("正在获取章节内容")
        for i in range(len(self.chapter_url)):
            response1 = requests.get(self.chapter_url[i])
            try:
                html1 = response1.text.encode('ISO-8859-1').decode('utf8')
                self.chapter_content.append(re.findall('<div id="content">(.*?)</div>', html1, re.S)[0])
                #调试输出内容，必要可注释
                print("已获取第%s章内容"%self.chapter_num+re.findall('<div id="content">(.*?)</div>', html1, re.S)[0])
            except Exception as e:
                print("第%s章节内容为空,无法获取"%self.chapter_num)
            self.chapter_num=self.chapter_num+1
            #每隔5秒获取一次章节内容
            # sleep(5)
        return self.chapter_url, self.chapter_name, self.chapter_content

    # 保存数据,需要传入小说名称,excel文件名称
    def SaveNovelDataToExcel(self,NovelName,ExcelName):
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename='%s.xlsx' % ExcelName)
        except Exception as e:
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.get_sheet_names()
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            wb.remove_sheet(wb.get_sheet_by_name(name))
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title = NovelName
        ws.cell(row=1, column=1, value='章节名称')
        ws.cell(row=1, column=2, value='章节地址')
        ws.cell(row=1, column=3, value='章节内容')
        ws.cell(row=1, column=4, value='小说作者')
        ws.cell(row=2, column=4, value=self.author[0])
        ws.cell(row=1, column=5, value='小说状态')
        ws.cell(row=2, column=5, value=self.status[0])
        ws.cell(row=1, column=6, value='最后更新时间')
        ws.cell(row=2, column=6, value=self.last_update_time[0])
        ws.cell(row=1, column=7, value='最新一章')
        ws.cell(row=2, column=7, value=self.last_chapter_name[0])
        for i in range(len(self.chapter_name)):
            ws.cell(row=i + 2, column=1, value=self.chapter_name[i])
        for i in range(len(self.chapter_url)):
            ws.cell(row=i + 2, column=2, value=self.chapter_url[i])
        for i in range(len(self.chapter_content)):
            ws.cell(row=i + 2, column=3, value=self.chapter_content[i])
        wb.save("%s.xlsx" % ExcelName)

if __name__=="__main__":
    print("正在实例化对象")
    Spider=Spider_bqg_novel()
    print("正在获取章节信息")
    Spider.GetChapterData("https://www.xxbiquge.com/79_79590/")
    print(len(Spider.chapter_content),len(Spider.chapter_name),len(Spider.chapter_url))
    print("正在存储数据")
    Spider.SaveNovelDataToExcel("剑镇鸿蒙","笔趣阁小说")