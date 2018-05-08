#coding=utf8
_Author_ = 'Alvis'
_Date_ = '2018-04-22 21:08'
_readme_="获取斗鱼直播信息"

import requests
import re
import json
import openpyxl
import time

#数据处理
class CategoryData:

    #初始化分类存储列表
    def __init__(self):
        #祖分类信息
        self.Ancestor_CategoryInfo=[]
        #祖分类名称
        self.Ancestor_Category_Name=[]
        #祖分类URL
        self.Ancestor_Category_Url=[]
        #父分类信息
        self.Parent_CategoryInfo=[]
        # 父分类名称
        self.Parent_Category_Name = []
        # 父分类URL
        self.Parent_Category_Url = []
        #子分类标签
        self.Sub_CategoryInfo = []
        #主播信息
        self.Anchor_Info=[]

    #获取祖分类
    def Get_Ancestor_Category_Data(self):

        # 总分类url地址
        url = "https://www.douyu.com/directory"

        try:

            # 获取页面信息内容
            response = requests.get(url)
            html = response.text

            # 获取关键信息，包含所有的大分类囊括信息
            content1 = re.findall('<div class="classify-li">(.*?)</div>', html, re.S)[0]
            content2 = re.findall('<a(.*?)</li>', content1, re.S)

            # 截取关键信息
            for i in content2:
                Ancestor_Category_name = re.findall('>(.*?)</a>', i, re.S)[0].strip()
                Ancestor_Category_url = "https://www.douyu.com" + re.findall('data-href="(.*?)"', i, re.S)[0]
                self.Ancestor_CategoryInfo.append([Ancestor_Category_name, Ancestor_Category_url])

            for info in self.Ancestor_CategoryInfo:
                # 祖分类名称
                self.Ancestor_Category_Name.append(info[0])
                # 祖分类URL
                self.Ancestor_Category_Name.append(info[1])

            with open("douyu.log","a+") as f:
                f.write("已获取祖分类数据,一共有%s个祖分类 "%len(self.Ancestor_CategoryInfo)+time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        except Exception as re_error:
            with open("douyu.log","a+") as f:
                f.write("页面无法相应,错误:%s  "%re_error+time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        return self.Parent_CategoryInfo

    #保存祖分类
    def Save_Ancestor_Category_Data(self):
        with open("douyu.log", 'a+') as f:
            f.write("检查获取祖分类函数返回的祖分类数据为%s  "%len(self.Ancestor_CategoryInfo) + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            f.close()
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename="斗鱼.xlsx")
        except Exception as e:
            with open("douyu.log", 'a+') as f:
                f.write("斗鱼.xlsx文件名不存在,进行创建  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                f.close()
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title ='斗鱼祖分类'
        ws.cell(row=1, column=1, value='祖分类名称')
        ws.cell(row=1, column=2, value='祖分类Url')
        for i in range(len(self.Ancestor_CategoryInfo)):
            try:
                ws.cell(row=i+2, column=1, value=self.Ancestor_CategoryInfo[i][0])
                ws.cell(row=i+2, column=2, value=self.Ancestor_CategoryInfo[i][1])
            except Exception as BigCategoryInfo_error:
                with open("douyu.log", 'a+') as f:
                    f.write("无法写入数据,错误:%s  "%BigCategoryInfo_error + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                    f.close()
        wb.save("斗鱼.xlsx")

    #获取父分类
    def Get_Parent_Category_Data(self):
        for i in range(1,len(self.Ancestor_CategoryInfo)):
            #获取祖分类的名称及url地址
            url_ancestor_category=self.Ancestor_CategoryInfo[i][1]
            name_ancestor_category=self.Ancestor_CategoryInfo[i][0]
            #获取各个总分类的页面源代码
            html=requests.get(url_ancestor_category).text
            #获取各个祖分类下的父分类所在数据区
            contents=re.findall('<ul id="live-list-contentbox">(.*?)</ul>',html,re.S)[0]
            #获取所有的父分类名称
            name_parent_category=re.findall('<p class="title">(.*?)</p>',contents,re.S)
            #获取所有的父分类url
            url=re.findall('href="(.*?)"',contents,re.S)
            #为url添加https://www.douyu.com
            url_parent_category=[]
            for j in range(len(url)):
                url_parent_category.append("https://www.douyu.com" + url[j])
            #所有父分类信息
            for k in range(len(url_parent_category)):
                self.Parent_CategoryInfo.append([name_ancestor_category,url_ancestor_category,name_parent_category[k],url_parent_category[k]])

        for info in self.Parent_CategoryInfo:
            # 父分类名称
            self.Parent_Category_Name.append(info[2])
            # 父分URL
            self.Parent_Category_Url.append(info[3])

    #保存父分类
    def Save_Parent_Category_Data(self):
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename="斗鱼.xlsx")
        except Exception as e:
            with open("douyu.log", 'a+') as f:
                f.write("斗鱼.xlsx文件名不存在,进行创建  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                f.close()
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title ='斗鱼父分类'
        ws.cell(row=1, column=1, value='祖分类名称')
        ws.cell(row=1, column=2, value='祖分类Url')
        ws.cell(row=1, column=3, value='父分类名称')
        ws.cell(row=1, column=4, value='父分类Url')
        for i in range(len(self.Parent_CategoryInfo)):
            try:
                ws.cell(row=i + 2, column=1, value=self.Parent_CategoryInfo[i][0])
                ws.cell(row=i + 2, column=2, value=self.Parent_CategoryInfo[i][1])
                ws.cell(row=i + 2, column=3, value=self.Parent_CategoryInfo[i][2])
                ws.cell(row=i + 2, column=4, value=self.Parent_CategoryInfo[i][3])
            except Exception as ParentCategoryInfo_error:
                with open("douyu.log", 'a+') as f:
                    f.write("无法写入数据,错误:%s  "%ParentCategoryInfo_error + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                    f.close()
        wb.save("斗鱼.xlsx")

    #获取子分类
    def Get_Sub_Category_Data(self):
        for i in range(len(self.Parent_CategoryInfo)):
            response=requests.get(self.Parent_CategoryInfo[i][3])
            html=response.text
            content=re.findall('data-live-list-type="(.*?)"',html,re.S)
            if content==[]:
                content2 = "无子分类标签"
                self.Sub_CategoryInfo.append([self.Parent_CategoryInfo[i], content2])
            else:
                del content[0]
                self.Sub_CategoryInfo.append([self.Parent_CategoryInfo[i], str(content).strip('[').strip(']').replace("'","")])

    #保存子分类
    def Save_Sub_Category_Data(self):
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename="斗鱼.xlsx")
        except Exception as e:
            with open("douyu.log", 'a+') as f:
                f.write("斗鱼.xlsx文件名不存在,进行创建  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                f.close()
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title ='斗鱼子分类'
        ws.cell(row=1, column=1, value='祖分类名称')
        ws.cell(row=1, column=2, value='祖分类Url')
        ws.cell(row=1, column=3, value='父分类名称')
        ws.cell(row=1, column=4, value='父分类Url')
        ws.cell(row=1, column=5, value='子分类标签')
        for i in range(len(self.Sub_CategoryInfo)):
            try:
                ws.cell(row=i + 2, column=1, value=self.Sub_CategoryInfo[i][0][0])
                ws.cell(row=i + 2, column=2, value=self.Sub_CategoryInfo[i][0][1])
                ws.cell(row=i + 2, column=3, value=self.Sub_CategoryInfo[i][0][2])
                ws.cell(row=i + 2, column=4, value=self.Sub_CategoryInfo[i][0][3])
                ws.cell(row=i + 2, column=5, value=self.Sub_CategoryInfo[i][1])
            except Exception as SubCategoryInfo_error:
                with open("douyu.log", 'a+') as f:
                    f.write("无法写入数据,错误:%s  "%SubCategoryInfo_error + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                    f.close()
        wb.save("斗鱼.xlsx")

    #判断是否存在祖父分类
    def Is_Exist_Category(self,ancestor,parent):

        #判断是否存在祖分类
        global result
        result=[]
        if ancestor in self.Ancestor_Category_Name:
            #判断是否存在父分类
            if parent in self.Parent_Category_Name:
                for info in self.Parent_CategoryInfo:
                    if ancestor in info and parent in info:
                        with open("douyu.log", "a+") as f:
                            f.write("祖父分类都存在  "+time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                        result=info
                        break
            else:
                result=False
                with open("douyu.log","a+") as f:
                    f.write("父分类不存在  "+time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        else:
            result=False
            with open("douyu.log", "a+") as f:
                f.write("祖分类不存在  "+time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

        return result

    #根据提供的祖父分类及页数，批量爬取该分类的主播信息
    def Get_Anchor_Info(self,ancestor,parent,type_num,page):

        result=self.Is_Exist_Category(ancestor,parent)

        if result is False:
            with open("douyu.log", "a+") as f:
                f.write("分类不存在,无法获取主播信息  "+time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        else:
            with open("douyu.log", "a+") as f:
                f.write("分类存在,准备获取主播信息  "+time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

            url=result[3]
            response=requests.get(url)
            html=response.text
            l=re.findall('<a class="play-list-link"',html,re.S)

            if len(l)==120:

                for page_num in range(1,page+1):
                    a='https://www.douyu.com/gapi/rkc/directory/2_'
                    api=a+"%s"%type_num+"/"+"%s"%(page_num)
                    response2=requests.get(api)
                    html2=response2.text
                    content=json.loads(html2)
                    #收集每一页的主播信息
                    for i in range(len(content['data']['rl'])):
                        #作用数据区
                        data=content['data']['rl'][i]
                        #主播名称
                        name=data['nn']
                        #主播房间标题
                        title=data['rn']
                        #主播所在父分类
                        parent_category=data['c2name']
                        #主播房间链接
                        room="https://www.douyu.com"+data['url']
                        #主播所在房间现时人数
                        hot=data['ol']
                        #主播标签
                        h=[]
                        try:
                            if len(data['utag']) >1:
                                for s in data['utag']:
                                    h.append(s['name'])
                                tag=h
                        except Exception:
                            tag="主播无标签"
                        finally:
                            self.Anchor_Info.append([name,title,parent_category,room,hot,tag])

                for i in range(len(self.Anchor_Info)):
                    print(i+1,self.Anchor_Info[i])

            else:

                # 获取主播名称
                anchor_name = []
                name = re.findall('<span class="dy-name ellipsis fl">(.*?)</span>', html, re.S)
                for content in name:
                    anchor_name.append(content)

                #获取主播房间标题
                anchor_title = []
                title=re.findall('<h3 class="ellipsis">(.*?)</h3>',html,re.S)
                for content in title:
                    anchor_title.append(content.strip())

                #获取主播父分类
                anchor_parent_category = []
                parent_category=re.findall('<span class="tag ellipsis">(.*?)</span>',html,re.S)
                for content in parent_category:
                    anchor_parent_category.append(content)

                #获取主播房间链接
                anchor_room = []
                room=re.findall('data-sub_rt="0" href="(.*?)"',html,re.S)
                for content in room:
                    anchor_room.append("https://www.douyu.com"+content)

                #获取主播现时人数
                anchor_hot = []
                hot=re.findall('<span class="dy-num fr"  >(.*?)</span>',html,re.S)
                for content in hot:
                    anchor_hot.append(content)

                for i in range(len(anchor_name)):
                    self.Anchor_Info.append([anchor_name[i],anchor_title[i],anchor_parent_category[i],
                                       anchor_room[i],anchor_hot[i]])

                for data in self.Anchor_Info:
                    print(data)






#主函数运行
if __name__=="__main__":
    #实例化
    Spider = CategoryData()
    Spider.Get_Ancestor_Category_Data()
    Spider.Get_Parent_Category_Data()
    Spider.Get_Anchor_Info('网游竞技','传奇',1,1)
