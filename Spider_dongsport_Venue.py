#coding=utf8
_Author_ = 'Alvis'
_Date_ = '14:40 2018-05-04'

import requests
import re
from requests.packages import urllib3
import openpyxl

#球馆名称
venue_name=[]
#球馆地址
venue_address=[]
#球馆电话
venue_mobile=[]
#球馆标签
venue_tag=[]
#球馆网址
venue_web_address=[]
#城市信息
city_info=[]
#城市名称
city_name=[]
#城市url地址
city_url=[]
#最大也数
max_page = []
#第一页地址
first_page=[]

def GetCityInfo():
    dw_url='http://www.dongsport.com/'
    response=requests.get(dw_url)
    html=response.text
    city_url_contents=re.findall('<li><a href="(.*?)"',re.findall('<ul class="cityUL">(.*?)</ul>',html,re.S)[0],re.S)
    city_name=re.findall('rel="nofollow">(.*?)<',re.findall('<ul class="cityUL">(.*?)</ul>',html,re.S)[0],re.S)
    first_page_contents = []

    for url in city_url_contents:
        #获取每个城市场馆列表的第一页
        response1=requests.get("http://www.dongsport.com"+url)
        html1=response1.text
        cookies=response1.cookies
        #携带cookies访问venue,获取最大页数
        dw_url2='http://www.dongsport.com/venue/'
        response2=requests.get(dw_url2,cookies=cookies)
        html2=response2.text
        max_page.append(re.findall('<span style="display:block; width:42px; height:17px;line-height:17px;">1/(.*?)<',html2,re.S)[0])
        first_page_contents.append(re.findall('venue/list-(.*?)1.html', html2, re.S)[0])

    for url in city_url_contents:
        city_url.append("http://www.dongsport.com"+url)
    for page in first_page_contents:
        first_page.append("http://www.dongsport.com/venue/list-"+page)
    for i in range(len(city_name)):
        city_info.append((city_name[i],"http://www.dongsport.com"+city_url_contents[i],max_page[i],first_page[i]))

def GetVenueData():
    for info in city_info:
        print("正在爬取%s的场馆信息"%info[0])
        for page in range(1,int(info[2])+1):
            print("当前访问%s场馆第%s页"%(info[0],page))
            url=info[3]+"%s"%page+".html"
            try:
                response = requests.get(url)
                urllib3.disable_warnings()
            except Exception as error:
                with open("log.text", "a+") as f:
                    f.write("无法访问%s场馆第%s页" % (info[0],page) + "错误为:%s" % error + "\n")
                continue
            html = response.text
            venue_content = re.findall('<div class="left v_l_text">(.*?)</div>', html, re.S)
            for content in venue_content:
                venue_name.append(re.findall('target="_blank">(.*?) ', content, re.S)[0].strip())
                venue_address.append(re.findall('<li>(.*?) ', content, re.S)[0].strip())
                venue_mobile.append(re.findall('<b class="fontstyle4">(.*?) ', content, re.S)[0].strip())
                venue_tag.append(re.findall('<li>(.*?)</li>', content, re.S)[2].strip())
                venue_web_address.append("http://www.dongsport.com" + re.findall('href="(.*?)"', content, re.S)[0])
            print("已获取%s第%s页数据" % (info[0],page))

def SaveVenueData():
    try:
        wb=openpyxl.load_workbook(filename="场馆信息.xlsx")
    except Exception as e:
        wb=openpyxl.Workbook()
    # 获取所有的表
    all = wb.sheetnames
    # 删除表Sheet
    name = 'Sheet'
    if name in all:
        del wb['Sheet']
    ws=wb.create_sheet()
    ws.title='动网'
    ws.cell(row=1, column=1, value='场馆名称')
    ws.cell(row=1, column=2, value='场馆地址')
    ws.cell(row=1, column=3, value='场馆电话')
    ws.cell(row=1, column=4, value='场馆标签')
    ws.cell(row=1, column=5, value='场馆网址')
    for i in range(len(venue_mobile)):
        ws.cell(row=i + 2, column=1, value=venue_name[i])
        ws.cell(row=i + 2, column=2, value=venue_address[i])
        ws.cell(row=i + 2, column=3, value=venue_mobile[i])
        ws.cell(row=i + 2, column=4, value=venue_tag[i])
        ws.cell(row=i + 2, column=5, value=venue_web_address[i])
        print("已写入%s条数据"%(i+1))
    wb.save("场馆信息.xlsx")

def Get_One_VenueInfo():
    for page in range(1,410):
        print("当前访问第%s页"%page)
        url='http://www.dongsport.com/venue/list-1004401-0-0-0-0-0-0-0-%s.html'%page
        try:
            response=requests.get(url)
            urllib3.disable_warnings()
        except Exception as error:
            with open("log.text","a+") as f:
                f.write("无法访问第%s页"%page+"错误为:%s"%error+"\n")
            continue
        html = response.text
        venue_content=re.findall('<div class="left v_l_text">(.*?)</div>',html,re.S)
        for content in venue_content:
            venue_name.append(re.findall('target="_blank">(.*?) ',content,re.S)[0].strip())
            venue_address.append(re.findall('<li>(.*?) ',content,re.S)[0].strip())
            venue_mobile.append(re.findall('<b class="fontstyle4">(.*?) ',content,re.S)[0].strip())
            venue_tag.append(re.findall('<li>(.*?)</li>', content, re.S)[2].strip())
            venue_web_address.append("http://www.dongsport.com"+re.findall('href="(.*?)"',content,re.S)[0])
        print("已获取第%s页数据"%page)

def Save_One_VenueInfo():
    try:
        wb=openpyxl.load_workbook(filename="场馆信息.xlsx")
    except Exception as e:
        wb=openpyxl.Workbook()
    # 获取所有的表
    all = wb.sheetnames
    # 删除表Sheet
    name = 'Sheet'
    if name in all:
        del wb['Sheet']
    ws=wb.create_sheet()
    ws.title='动网'
    ws.cell(row=1, column=1, value='场馆名称')
    ws.cell(row=1, column=2, value='场馆地址')
    ws.cell(row=1, column=3, value='场馆电话')
    ws.cell(row=1, column=4, value='场馆标签')
    ws.cell(row=1, column=5, value='场馆网址')
    for i in range(len(venue_mobile)):
        ws.cell(row=i + 2, column=1, value=venue_name[i])
        ws.cell(row=i + 2, column=2, value=venue_address[i])
        ws.cell(row=i + 2, column=3, value=venue_mobile[i])
        ws.cell(row=i + 2, column=4, value=venue_tag[i])
        ws.cell(row=i + 2, column=5, value=venue_web_address[i])
        print("已写入%s条数据"%(i+1))
    wb.save("场馆信息.xlsx")


def run():
    GetCityInfo()
    print(city_info)
    GetVenueData()
    SaveVenueData()

if __name__=="__main__":
    run()