#coding=utf8
_Author_ = 'Alvis'
_Date_ = '2018-05-02 19:21'

import requests
import re
import openpyxl
from requests.packages import urllib3
from time import sleep

class Spider_bqg_novel:

    #初始化数据
    def __init__(self):
        #章节名列表
        self.chapter_name=[]
        #章节链接列表
        self.chapter_url=[]
        #章节内容
        self.chapter_content=[]
        #小说作者
        self.author=[]
        #小说状态
        self.status=[]
        #小说最后更新时间
        self.last_update_time=[]
        #小说最近更新的章节名称
        self.last_chapter_name=[]
        self.chapter_num=1

    #获取章节名称及链接、内容,需要传入小说页面链接
    def GetChapterData(self,url):
        # 返回信息
        urllib3.disable_warnings()
        try:
            print("正在测试网址是否响应")
            response = requests.get(url,verify=False,timeout=5)
            if response.status_code==200:
                print("网址成功响应",response.status_code)
                # 源网页代码
                html = response.text.encode('ISO-8859-1').decode('utf8')

                # 获取小说作者
                try:
                    self.author = re.findall('<meta property="og:novel:author" content="(.*?)"/>', html, re.S)
                    print("已获取小说作者")
                except Exception as reason_author:
                    print("无法获取小说作者", reason_author)

                # 小说状态
                try:
                    self.status = re.findall('<meta property="og:novel:status" content="(.*?)"/>', html, re.S)
                    print("已获取小说状态")
                except Exception as reason_status:
                    print("无法获取小说状态",reason_status)

                # 获取最后更新时间
                try:
                    self.last_update_time = re.findall('<meta property="og:novel:update_time" content="(.*?)"/>', html, re.S)
                    print("已获取小说最后更新时间")
                except Exception as reason_last_update_time:
                    print("无法获取小说最后更新时间",reason_last_update_time)

                # 获取最近更新的章节名称
                try:
                    self.last_chapter_name = re.findall('<meta property="og:novel:latest_chapter_name" content="(.*?)"/>', html,re.S)
                    print("已获取最近更新的章节名称")
                except Exception as reason_last_chapter_name:
                    print("无法获取最近更新的章节名称",reason_last_chapter_name)


                # 网页中包含章节名称及链接的数据
                try:
                    content = re.findall('<dl>(.*?)</dl>', html, re.S)

                    try:
                        # 获取章节链接，缺少"https://www.xxbiquge.com"
                        print("正在获取章节链接")
                        chapter_url_content = re.findall('<dd><a href="(.*?)">', content[0], re.S)
                        print("已获取章节链接")
                        # 补全章节链接，补全"https://www.xxbiquge.com"
                        for link in chapter_url_content:
                            self.chapter_url.append("https://www.xxbiquge.com" + link)
                    except Exception as reason_chapter_url_content:
                        print("无法获取章节链接",reason_chapter_url_content)

                    try:
                        # 获取章节名称
                        print("正在获取章节名称")
                        self.chapter_name = re.findall('">(.*?)</a>', content[0], re.S)
                        print("已获取章节名称")
                    except Exception as reason_chapter_name:
                        print("无法获取章节名称",reason_chapter_name)

                except Exception as reason_chapter_url_content:
                    print("无法获取网页中包含章节名称及链接的数据",reason_chapter_url_content)


                # 获取章节内容
                print("正在获取章节内容")
                for i in range(len(self.chapter_url)):
                    response1 = requests.get(self.chapter_url[i])
                    try:
                        html1 = response1.text.encode('ISO-8859-1').decode('utf8')
                        self.chapter_content.append(re.findall('<div id="content">(.*?)</div>', html1, re.S)[0])
                        # 调试输出内容，必要可注释
                        print("已获取第%s章内容" % self.chapter_num + re.findall('<div id="content">(.*?)</div>', html1, re.S)[0])
                    except Exception as e:
                        print("第%s章节内容为空,无法获取" % self.chapter_num)
                        print(e)
                        continue
                    self.chapter_num = self.chapter_num + 1
                    # 每隔5秒获取一次章节内容
                    # sleep(5)

            else:
                print("网址响应错误")
                print(response.status_code)

        except Exception as a:
            print("网址无法相应",a)


    # 保存数据,需要传入小说名称,excel文件名称
    def SaveNovelDataToExcel(self,NovelName,ExcelName):
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename='%s.xlsx' % ExcelName)
        except Exception as e:
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        print(wb.sheetnames)
        # 为新表命名
        ws.title = NovelName
        
        try:
            ws.cell(row=1, column=1, value='小说作者')
            ws.cell(row=2, column=1, value=self.author[0])
        except Exception as write_false_author:
            print("无法写入作者",write_false_author)
        
        try:
            ws.cell(row=1, column=2, value='小说状态')
            ws.cell(row=2, column=2, value=self.status[0])
        except Exception as write_false_status:
            print("无法写入小说状态",write_false_status)
        
        try:
            ws.cell(row=1, column=3, value='最后更新时间')
            ws.cell(row=2, column=3, value=self.last_update_time[0])
        except Exception as write_false_last_update_time:
            print("无法写入小说最后更新时间",write_false_last_update_time)
        
        try:
            ws.cell(row=1, column=4, value='最新一章')
            ws.cell(row=2, column=4, value=self.last_chapter_name[0])
        except Exception as write_false_last_chapter_name:
            print("无法写入最新章节名称",write_false_last_chapter_name)
            
        ws.cell(row=1, column=5, value='章节名称')
        ws.cell(row=1, column=6, value='章节地址')
        ws.cell(row=1, column=7, value='章节内容')
        
        #写入章节名称
        for i in range(len(self.chapter_name)):
            try:
                ws.cell(row=i + 2, column=5, value=self.chapter_name[i])
            except Exception as write_false_chapter_name:
                print("无法写入该章节名称",write_false_chapter_name)
                continue
                
        #写入章节URL
        for i in range(len(self.chapter_url)):
            try:
                ws.cell(row=i + 2, column=6, value=self.chapter_url[i])
            except Exception as write_false_chapter_url:
                print("无法写入该章节Url",write_false_chapter_url)
                continue

        #写入章节内容
        for i in range(len(self.chapter_content)):
            try:
                ws.cell(row=i + 2, column=7, value=self.chapter_content[i])
            except Exception as write_false_chapter_content:
                print("无法写入改章节内容",write_false_chapter_content)
                continue

        wb.save("%s.xlsx" % ExcelName)

if __name__=="__main__":
    print("正在实例化对象")
    Spider=Spider_bqg_novel()
    print("正在获取章节信息")
    Spider.GetChapterData("https://www.xxbiquge.com/80_80269/")
    print(len(Spider.chapter_content),len(Spider.chapter_name),len(Spider.chapter_url))
    print("正在存储数据")
    Spider.SaveNovelDataToExcel("从垃圾工到星空战神","笔趣阁小说")