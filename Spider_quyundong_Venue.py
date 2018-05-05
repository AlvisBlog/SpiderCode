#coding=utf8
import requests
import openpyxl
import re
import json
class Quyundong:

    def __init__(self):
        self.venues_info=[]
        self.city_info=[]
        self.city_id=[]

    def GetCityInfo(self):
        url="http://www.quyundong.com"
        response=requests.get(url)
        html=response.text
        city_id=re.findall('data-cityId="(.*?)"',html,re.S)
        self.city_id=city_id
        city_name=re.findall('data-cityName="(.*?)"',html,re.S)
        page_num=[]
        for id in city_id:
            url='http://www.quyundong.com/?city_id=%s'%id
            response1=requests.get(url)
            html1=response1.text
            page_num.append(re.findall("data-count='(.*?)'",html1,re.S)[0])
        for i in range(len(city_id)):
            self.city_info.append((city_name[i],"http://www.quyundong.com/?city_id="+city_id[i],page_num[i]))

    def GetVenuesInfo(self):
        for j in range(len(self.city_info)):
            print("正在爬取%s城市的场馆信息"%self.city_info[j][0])
            for page in range(1,int(self.city_info[j][2])+1):
                api='http://www.quyundong.com/index/businesslist?random=0.5960742008869808&page=%s&city_id=%s'\
                    %(page,self.city_id[j])
                data=requests.get(api).text
                content=json.loads(data)
                for i in range(len(content['data']['data'])):
                    self.venues_info.append((content['data']['data'][i]['name'],content['data']['data'][i]['address'],
                                             content['data']['data'][i]['promote_price'],content['data']['data'][i]['comment_avg'],
                                        content['data']['data'][i]['comment_count'],content['data']['data'][i]['price'],
                                    content['data']['data'][i]['latitude'],content['data']['data'][i]['longitude']))
            print("已爬取%s城市的场馆信息" % self.city_info[j][0])

    def SaveVenuesInfo(self):
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename="场馆信息.xlsx")
        except Exception as e:
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title ='趣运动'
        ws.cell(row=1, column=1, value='场馆名')
        ws.cell(row=1, column=2, value='场馆地址')
        ws.cell(row=1, column=3, value='趣运动价格')
        ws.cell(row=1, column=4, value='球馆星级')
        ws.cell(row=1, column=5, value='评论数')
        ws.cell(row=1, column=6, value='其他价格')
        ws.cell(row=1, column=7, value='球馆纬度')
        ws.cell(row=1, column=8, value='球馆经度')

        for i in range(len(self.venues_info)):
            ws.cell(row=i + 2, column=1, value=self.venues_info[i][0])
            ws.cell(row=i + 2, column=2, value=self.venues_info[i][1])
            ws.cell(row=i + 2, column=3, value=self.venues_info[i][2])
            ws.cell(row=i + 2, column=4, value=self.venues_info[i][3])
            ws.cell(row=i + 2, column=5, value=self.venues_info[i][4])
            ws.cell(row=i + 2, column=6, value=self.venues_info[i][5])
            ws.cell(row=i + 2, column=7, value=self.venues_info[i][6])
            ws.cell(row=i + 2, column=8, value=self.venues_info[i][7])
            print("已写入%s个小数据" % i)

        wb.save("场馆信息.xlsx")


if __name__=='__main__':
    Spider=Quyundong()
    Spider.GetCityInfo()
    Spider.GetVenuesInfo()
    Spider.SaveVenuesInfo()
